from openpyxl import load_workbook
from xlsx_import.group_hierarchy import Lesson, Group
from xlsx_import.functions import contains_alphabet_character


def get_matrix(sheet, min_row: int, max_row: int, min_col: int, max_col: int):
    sheet_matrix = []
    for row in sheet.iter_rows(
            min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        row_values = []
        for cell in row:
            row_values.append(None if cell.value == "" else cell.value)  # если клетка не NoneType, но пустая, то замена
        sheet_matrix.append(row_values)
    return sheet_matrix


def parse_group_schedule(schedule: list[list]) -> list[Lesson]:
    group_lessons = []
    day = 1
    lesson_count = 1
    for lesson in schedule:
        if lesson_count == 15:
            lesson_count = 1
            day += 1
        if contains_alphabet_character(lesson[0]):
            parsed_lesson = Lesson.from_list(
                lesson_num=lesson_count,
                day_number=day,
                classroom=lesson[3],
                teacher_name=lesson[2],
                lesson_type=lesson[1],
                discipline=lesson[0]
            )
            group_lessons.append(parsed_lesson)
        lesson_count += 1
    return group_lessons


def addition_by_sheet_type(sheet_type: int, switch: bool):
    """adjusts start position of the matrix that will be taken by get_matrix function"""
    if sheet_type == 0:
        return 9
    if sheet_type == 1:
        return 10
    if sheet_type == 2:
        if switch is False:
            return 5
        if switch is True:
            return 10
    if sheet_type == 3:
        return 4
    if sheet_type == 4:
        if switch is False:
            return 9
        if switch is True:
            return 10


def get_multiple_schedules(sheet, matrices_num: int, sheet_type: int) -> list[Group]:
    schedules = []
    min_column = 6  # start offset

    switch = False
    for i in range(matrices_num):
        max_column = min_column + 3  # end offset
        matrix = get_matrix(sheet, 4, 87, min_column, max_column)
        min_column += addition_by_sheet_type(sheet_type, switch)
        # adding to min_column number to adjust matrix start position

        switch = not switch
        group_timetable = parse_group_schedule(matrix)  # get schedule of each group from one sheet

        schedules.append(group_timetable)
    prefix_list = get_year_prefixes(sheet, matrices_num, sheet_type)  # get name of each group from one sheet
    group_list = []
    for i in range(len(schedules)):
        group_list.append(
            Group(
                group_name=prefix_list[i],
                schedule=schedules[i]
            ))  # create Group and append to schedules

    return group_list


def get_group_prefix(working_sheet, column: int) -> str:
    return working_sheet.cell(row=2, column=column).value  # group prefix is always in second row, only column matters


def get_year_prefixes(working_sheet, amount_of_groups: int, sheet_type: int) -> list[str]:
    init_column = 6  # first position of every prefix in sheet
    switch = False
    prefixes = []
    for _ in range(amount_of_groups):
        prefixes.append(get_group_prefix(working_sheet, init_column))
        init_column += addition_by_sheet_type(sheet_type, switch)
        switch = not switch
    return prefixes


def read_schedule(filename: str):
    workbook = load_workbook(filename)
    return workbook


def get_groups_timetable(sheet_path: str) -> list[Group]:
    wb = read_schedule(sheet_path)
    sheet_names = wb.sheetnames

    complete_timetable = []

    for item in sheet_names:
        match item:
            case str('1 курс ' | '2 курс '):
                complete_timetable.extend(get_multiple_schedules(wb[item], 3, 0))
            case str('4 курс '):
                complete_timetable.extend(get_multiple_schedules(wb[item], 3, 1))
            case str('вечернее'):
                complete_timetable.extend(get_multiple_schedules(wb[item], 4, 2))
            case str('1 магистратура' | '2 магистратура'):
                complete_timetable.extend(get_multiple_schedules(wb[item], 2, 3))
            case str('3 курс '):
                complete_timetable.extend(get_multiple_schedules(wb[item], 3, 4))
    return complete_timetable


def groups_to_json(timetable, path):
    for idx, group_class in enumerate(timetable, 1):
        import json
        import cattrs

        structure = group_class.schedule
        unstructured = cattrs.unstructure(structure)
        json_str = json.dumps(unstructured)
        open(f"{path}/{group_class.group_name}.json", 'w').write(json_str)
